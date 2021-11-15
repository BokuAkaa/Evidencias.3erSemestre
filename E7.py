# Autores :
# ANA BARBARA SARMIENTO NAJERA
# FELIPE DE JESUS PEREZ LOPEZ
# GONZALO ESCAJEDA ANGEL
# E-MANUEL MANZANAREZ MORA

from pyhunter import PyHunter
# API de hunter.io
from openpyxl import Workbook
# Biblioteca de Python para leer/escribir archivos xlsx/xlsm/xltx/xltm.
import getpass
# Solicita al usuario una contraseña sin imprimirla en pantalla.


def Busqueda(organizacion):
    resultado = hunter.domain_search(
        company=organizacion, limit=1,
        emails_type='personal')
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.remove_sheet(libro.get_sheet_by_name("Sheet"))

    for valores in datosEncontrados["emails"]:
        for datos in valores:
            if datos == 'value':
                emaildato = valores[datos]
    for valores in datosEncontrados["emails"]:
        for datos in valores:
            if datos == 'type':
                tipodato = valores[datos]

    hoja.cell(1, 1, "dominio")
    hoja.cell(1, 2, "país")
    hoja.cell(1, 3, "organizacion")
    hoja.cell(1, 4, 'email')
    hoja.cell(1, 5, 'tipo de email')
    hoja.cell(2, 1, datosEncontrados["domain"])
    hoja.cell(2, 2, datosEncontrados["country"])
    hoja.cell(2, 3, datosEncontrados["organization"])
    hoja.cell(2, 4, emaildato)
    hoja.cell(2, 5, tipodato)
    libro.save("Hunter" + organizacion + ".xlsx")

# La función "GuardarInformacion" crea un archivo .xlsx
# y nombra las columnas "dominio","país","organizacion"
# luego en la siguiente fila escribe los datos obtenidos como corresponde.

# Inicio:muestra el mensaje y pide al usuario escribir
# su contraseña y algún dominio de interés.

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
# apikey solicita la clave de la API de hunter.io
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
# Solicita algún dominio de interés para el usuario
# y envía la entrada de la variable "orga" a
# la función "Busqueda".

if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(datosEncontrados["domain"])
    print(datosEncontrados["country"])
    print(datosEncontrados["organization"])
    for valores in datosEncontrados["emails"]:
        for datos in valores:
            if datos == 'value':
                print(valores[datos])
    for valores in datosEncontrados["emails"]:
        for datos in valores:
            if datos == 'type':
                print(valores[datos])

    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)

# En caso de no encontrar nada el programa se cierra automaticamente
# En caso de encontrar algo imprime los resultados obtenidos
# almacenados en la variable "datosEncontrados"

# instalar modulos necesarios:
# pip install pyhunter
# pip install openpyxl
